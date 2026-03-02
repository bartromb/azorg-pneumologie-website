#!/usr/bin/env python3
"""
build.py — Genereer de Pneumologie-website vanuit Excel-bestanden (NL + FR + EN)
===============================================================================
Gebruik:
    python build.py              # genereert NL + FR + EN
    python build.py --lang nl    # alleen Nederlands
    python build.py --lang fr    # alleen Frans
    python build.py --lang en    # alleen Engels

Vereisten:
    pip install openpyxl jinja2 deep-translator

Vertaling:
    - Nieuwe NL-tekst → automatisch vertaald via deep-translator (Google)
    - Vertaling wordt opgeslagen in data/translations.json (cache)
    - Zonder internet: valt terug op cache, onvertaalde tekst = NL
"""
import sys, json, hashlib
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from jinja2 import Environment, BaseLoader
except ImportError:
    print("❌ pip install openpyxl jinja2"); sys.exit(1)

BASE = Path(__file__).parent
DATA_DIR = BASE / "data"
OUTPUT_DIR = BASE / "output"
CACHE_FILE = DATA_DIR / "translations.json"

COLOR_MAP = {
    "rood": "var(--red)", "teal": "var(--teal)", "groen": "var(--green)",
    "geel": "var(--yellow)", "paars": "var(--purple)", "donkerrood": "var(--dark)",
}
CAMPUS_CC = {
    "rood": "", "teal": "teal", "groen": "green",
    "geel": "yellow", "paars": "purple", "donkerrood": "dark",
}
PARTNERS = {
    "nl": ["Thoraxchirurgie", "Radiologie", "Oncologie", "Radiotherapie",
           "Cardiologie", "Pathologische Anatomie", "Intensieve Zorgen",
           "Kinesitherapie", "Psychologie"],
    "fr": ["Chirurgie thoracique", "Radiologie", "Oncologie", "Radiothérapie",
           "Cardiologie", "Anatomie pathologique", "Soins intensifs",
           "Kinésithérapie", "Psychologie"],
    "en": ["Thoracic Surgery", "Radiology", "Oncology", "Radiotherapy",
           "Cardiology", "Pathological Anatomy", "Intensive Care",
           "Physiotherapy", "Psychology"],
    "de": ["Thoraxchirurgie", "Radiologie", "Onkologie", "Strahlentherapie",
           "Kardiologie", "Pathologische Anatomie", "Intensivmedizin",
           "Physiotherapie", "Psychologie"],
}

# ─── UI strings (hardcoded, not from Excel) ──────────────────────────
UI = {
  "nl": {
    "lang":"nl",
    "page_title":"Pneumologie / Longziekten — AZORG Ziekenhuis",
    "meta_desc":"Dienst Pneumologie van AZORG. {n} longartsen op {c} campussen.",
    "skip":"Ga naar inhoud","back":"← AZORG Ziekenhuis","news":"Nieuws",
    "portal":"Patiëntenportaal (mynexuzhealth)",
    "patient":"Patiënt","pro":"Professional","praktisch":"Praktisch",
    "sub":"Longziekten",
    "n_over":"Over de dienst","n_camp":"Campussen","n_arts":"Artsen",
    "n_aand":"Aandoeningen","n_klin":"Klinieken","n_ha":"Huisartsen",
    "n_contact":"Contact","n_cta":"Online afspraak",
    "bc1":"AZORG","bc2":"Onze medische diensten",
    "bc3":"Pneumologie / Longziekten",
    "pills":"Actief op","btn_afs":"Online afspraak",
    "btn_sec":"Secretariaten &amp; campussen",
    "qa1":"Secretariaat bellen","qa2":"Online afspraak",
    "qa2s":"via mynexuzhealth","qa3":"Aandoeningen",
    "qa3s":"Longaandoeningen &amp; ziektebeelden",
    "qa4":"Mijn dossier","qa4s":"mynexuzhealth portaal",
    "l_over":"Over de dienst","l_camp":"Secretariaat per campus",
    "h_camp":"Contacteer ons op uw dichtstbijzijnde campus",
    "s_camp":"Maak een afspraak via telefoon, e-mail of online via mynexuzhealth. Alle campussen zijn bereikbaar van maandag tot en met vrijdag.",
    "camp_arts":"Longartsen op deze campus",
    "l_arts":"Ons team","h_arts":"Gespecialiseerde longartsen",
    "s_arts":"Alle longartsen zijn geconventioneerd. Elk heeft een brede basisexpertise aangevuld met deelspecialisaties.",
    "conv":"✓ Geconventioneerd",
    "arts_link":"Profiel &amp; foto op azorg.be",
    "btn_arts":"Bekijk alle artsen op azorg.be →",
    "l_aand":"Longaandoeningen","h_aand":"Aandoeningen die wij behandelen",
    "s_aand":'Meer informatie vindt u op <a href="https://www.azorg.be/nl/longziekten" style="color:var(--red);font-weight:600" target="_blank" rel="noopener noreferrer">azorg.be/longziekten</a>.',
    "meer":"Meer info →",
    "l_ond":"Diagnostiek &amp; ingrepen",
    "h_ond":"Longonderzoeken en -ingrepen",
    "s_ond":"De dienst beschikt over moderne diagnostische en interventionele technieken.",
    "l_klin":"Gespecialiseerde klinieken",
    "h_klin":"Expertklinieken binnen Pneumologie",
    "s_klin":"Naast de algemene consultaties biedt de afdeling gespecialiseerde klinieken.",
    "l_nws":"Nieuws","h_nws":"Actueel uit de dienst Pneumologie",
    "s_nws":"Recente ontwikkelingen, studies en nieuws over longziekten.",
    "btn_nws":"Bekijk al het nieuws op azorg.be →",
    "cta_btn":"Online afspraak maken →",
    "l_ha":"Voor professionals","h_ha":"Informatie voor huisartsen",
    "s_ha":"Snelle verwijzing, digitaal overleg en MOC-vergaderingen.",
    "ha1":"Nexuzhealth consult",
    "ha1d":"Digitaal overleg via het beveiligde nexuzhealth-platform.",
    "ha2":"MOC Longkanker",
    "ha2d":"Wekelijks multidisciplinair oncologisch overleg.",
    "ha3":"Laboratoriumgidsen","ha3d":"Toegang tot de laboratoriumgids.",
    "ha4":"Preoperatieve onderzoeken",
    "ha4d":"Richtlijnen voor preoperatieve evaluatie.",
    "l_ct":"Contact &amp; afspraak",
    "h_ct":"Afspraak maken bij Pneumologie",
    "h3_pr":"Praktische informatie",
    "on_desc":"afspraken maken kan 24/7 online via mynexuzhealth",
    "rook":"Rookstopkliniek","dring":"Dringende gevallen",
    "dring_d":"Longarts on-call 24/7 bereikbaar via de spoeddienst van AZORG",
    "h3_form":"Online aanvraag",
    "f_vn":"Voornaam","f_an":"Achternaam","f_em":"E-mailadres",
    "f_tel":"Telefoonnummer","f_camp":"Voorkeurscampus",
    "f_camp_ph":"Selecteer campus","f_red":"Reden van consultatie",
    "f_red_ph":"Selecteer categorie","f_info":"Bijkomende informatie",
    "f_info_ph":"Beschrijf uw klachten of vermeld uw verwijzende arts...",
    "f_submit":"Aanvraag versturen →",
    "f_thx":"Bedankt voor uw aanvraag.\\nOns secretariaat neemt binnen 2 werkdagen contact op.",
    "f_opts":["Kortademigheid / hoest","Slaapapneu / snurken",
              "Allergie / astma","Rookstop","Longkanker opvolging",
              "Longrevalidatie","Sportkliniek / duikkeuring",
              "Controle bestaande aandoening","Verwijzing huisarts","Andere"],
    "beh":"Wat behandelen wij?",
    "part":"Multidisciplinaire samenwerking met",
    "ft_tag":"Zorg van A tot Z",
    "ft_desc":"Pneumologie / Longziekten maakt deel uit van AZORG, het op twee na grootste ziekenhuis van Vlaanderen.",
    "ft_aand":"Aandoeningen","ft_klin":"Klinieken","ft_pr":"Praktisch",
    "ft_on":"Online afspraak","ft_dos":"Mijn dossier",
    "ft_pat":"Info voor patiënten","ft_pro":"Info voor professionals",
    "ft_copy":"© 2025 AZORG Ziekenhuis — Pneumologie / Longziekten",
  },
  "fr": {
    "lang":"fr",
    "page_title":"Pneumologie / Maladies pulmonaires — AZORG Hôpital",
    "meta_desc":"Service de Pneumologie d'AZORG. {n} pneumologues sur {c} campus.",
    "skip":"Aller au contenu","back":"← AZORG Hôpital","news":"Actualités",
    "portal":"Portail patient (mynexuzhealth)",
    "patient":"Patient","pro":"Professionnel","praktisch":"Pratique",
    "sub":"Maladies pulmonaires",
    "n_over":"Le service","n_camp":"Campus","n_arts":"Médecins",
    "n_aand":"Pathologies","n_klin":"Cliniques",
    "n_ha":"Médecins traitants","n_contact":"Contact",
    "n_cta":"Rendez-vous en ligne",
    "bc1":"AZORG","bc2":"Nos services médicaux",
    "bc3":"Pneumologie / Maladies pulmonaires",
    "pills":"Actif sur","btn_afs":"Rendez-vous en ligne",
    "btn_sec":"Secrétariats &amp; campus",
    "qa1":"Appeler le secrétariat","qa2":"Rendez-vous en ligne",
    "qa2s":"via mynexuzhealth","qa3":"Pathologies",
    "qa3s":"Affections pulmonaires",
    "qa4":"Mon dossier","qa4s":"portail mynexuzhealth",
    "l_over":"Le service","l_camp":"Secrétariat par campus",
    "h_camp":"Contactez-nous sur votre campus le plus proche",
    "s_camp":"Prenez rendez-vous par téléphone, e-mail ou en ligne via mynexuzhealth. Tous les campus sont accessibles du lundi au vendredi.",
    "camp_arts":"Pneumologues sur ce campus",
    "l_arts":"Notre équipe","h_arts":"Pneumologues spécialisés",
    "s_arts":"Tous les pneumologues sont conventionnés. Chacun dispose d'une expertise de base complétée par des sous-spécialisations.",
    "conv":"✓ Conventionné",
    "arts_link":"Profil &amp; photo sur azorg.be",
    "btn_arts":"Voir tous les médecins sur azorg.be →",
    "l_aand":"Affections pulmonaires",
    "h_aand":"Pathologies que nous traitons",
    "s_aand":'Plus d\'informations sur <a href="https://www.azorg.be/nl/longziekten" style="color:var(--red);font-weight:600" target="_blank" rel="noopener noreferrer">azorg.be/longziekten</a>.',
    "meer":"Plus d'info →",
    "l_ond":"Diagnostic &amp; interventions",
    "h_ond":"Examens et interventions pulmonaires",
    "s_ond":"Le service dispose de techniques diagnostiques et interventionnelles modernes.",
    "l_klin":"Cliniques spécialisées",
    "h_klin":"Cliniques d'expertise en Pneumologie",
    "s_klin":"En plus des consultations générales, le service offre des cliniques spécialisées.",
    "l_nws":"Actualités",
    "h_nws":"Actualités du service de Pneumologie",
    "s_nws":"Développements récents, études et nouvelles.",
    "btn_nws":"Voir toutes les actualités sur azorg.be →",
    "cta_btn":"Prendre rendez-vous en ligne →",
    "l_ha":"Pour les professionnels",
    "h_ha":"Informations pour les médecins traitants",
    "s_ha":"Référencement rapide, concertation numérique et réunions COM.",
    "ha1":"Nexuzhealth consult",
    "ha1d":"Concertation numérique via la plateforme sécurisée nexuzhealth.",
    "ha2":"COM Cancer du poumon",
    "ha2d":"Concertation oncologique multidisciplinaire hebdomadaire.",
    "ha3":"Guides de laboratoire","ha3d":"Accès au guide de laboratoire.",
    "ha4":"Examens préopératoires",
    "ha4d":"Directives pour l'évaluation préopératoire.",
    "l_ct":"Contact &amp; rendez-vous",
    "h_ct":"Prendre rendez-vous en Pneumologie",
    "h3_pr":"Informations pratiques",
    "on_desc":"prise de rendez-vous 24h/24 en ligne via mynexuzhealth",
    "rook":"Clinique de sevrage tabagique","dring":"Urgences",
    "dring_d":"Pneumologue de garde 24h/24 joignable via les urgences d'AZORG",
    "h3_form":"Demande en ligne",
    "f_vn":"Prénom","f_an":"Nom","f_em":"Adresse e-mail",
    "f_tel":"Numéro de téléphone","f_camp":"Campus de préférence",
    "f_camp_ph":"Sélectionnez un campus",
    "f_red":"Motif de consultation",
    "f_red_ph":"Sélectionnez une catégorie",
    "f_info":"Informations complémentaires",
    "f_info_ph":"Décrivez vos symptômes ou mentionnez votre médecin référent...",
    "f_submit":"Envoyer la demande →",
    "f_thx":"Merci pour votre demande.\\nNotre secrétariat vous contactera dans les 2 jours ouvrables.",
    "f_opts":["Essoufflement / toux","Apnée du sommeil / ronflements",
              "Allergie / asthme","Sevrage tabagique",
              "Suivi cancer du poumon","Réhabilitation pulmonaire",
              "Clinique du sport / aptitude plongée",
              "Contrôle affection existante",
              "Référence médecin traitant","Autre"],
    "beh":"Ce que nous traitons",
    "part":"Collaboration multidisciplinaire avec",
    "ft_tag":"Des soins de A à Z",
    "ft_desc":"La Pneumologie fait partie d'AZORG, le troisième plus grand hôpital de Flandre.",
    "ft_aand":"Pathologies","ft_klin":"Cliniques","ft_pr":"Pratique",
    "ft_on":"Rendez-vous en ligne","ft_dos":"Mon dossier",
    "ft_pat":"Info patients","ft_pro":"Info professionnels",
    "ft_copy":"© 2025 AZORG Hôpital — Pneumologie / Maladies pulmonaires",
  },
  "en": {
    "lang":"en",
    "page_title":"Pulmonology / Respiratory Medicine — AZORG Hospital",
    "meta_desc":"Department of Pulmonology at AZORG. {n} pulmonologists across {c} campuses.",
    "skip":"Skip to content","back":"← AZORG Hospital","news":"News",
    "portal":"Patient portal (mynexuzhealth)",
    "patient":"Patient","pro":"Professional","praktisch":"Practical",
    "sub":"Respiratory Medicine",
    "n_over":"About","n_camp":"Campuses","n_arts":"Physicians",
    "n_aand":"Conditions","n_klin":"Clinics",
    "n_ha":"Referring physicians","n_contact":"Contact",
    "n_cta":"Book appointment",
    "bc1":"AZORG","bc2":"Our medical services",
    "bc3":"Pulmonology / Respiratory Medicine",
    "pills":"Active at","btn_afs":"Book appointment",
    "btn_sec":"Secretariats &amp; campuses",
    "qa1":"Call secretariat","qa2":"Book online",
    "qa2s":"via mynexuzhealth","qa3":"Conditions",
    "qa3s":"Lung diseases &amp; conditions",
    "qa4":"My records","qa4s":"mynexuzhealth portal",
    "l_over":"About the department","l_camp":"Secretariat per campus",
    "h_camp":"Contact us at your nearest campus",
    "s_camp":"Book an appointment by phone, e-mail or online via mynexuzhealth. All campuses are reachable Monday through Friday.",
    "camp_arts":"Pulmonologists at this campus",
    "l_arts":"Our team","h_arts":"Specialised pulmonologists",
    "s_arts":"All pulmonologists are conventioned. Each has broad core expertise supplemented by sub-specialisations.",
    "conv":"✓ Conventioned",
    "arts_link":"Profile &amp; photo on azorg.be",
    "btn_arts":"View all physicians on azorg.be →",
    "l_aand":"Lung conditions","h_aand":"Conditions we treat",
    "s_aand":'More information at <a href="https://www.azorg.be/nl/longziekten" style="color:var(--red);font-weight:600" target="_blank" rel="noopener noreferrer">azorg.be/longziekten</a>.',
    "meer":"More info →",
    "l_ond":"Diagnostics &amp; procedures",
    "h_ond":"Pulmonary examinations and procedures",
    "s_ond":"The department has modern diagnostic and interventional techniques.",
    "l_klin":"Specialised clinics",
    "h_klin":"Expert clinics within Pulmonology",
    "s_klin":"Besides general consultations, the department offers specialised clinics.",
    "l_nws":"News","h_nws":"Latest from Pulmonology",
    "s_nws":"Recent developments, studies and respiratory medicine news.",
    "btn_nws":"View all news on azorg.be →",
    "cta_btn":"Book an appointment online →",
    "l_ha":"For professionals","h_ha":"Information for referring physicians",
    "s_ha":"Quick referral, digital consultation and MDT meetings.",
    "ha1":"Nexuzhealth consult",
    "ha1d":"Digital consultation via the secure nexuzhealth platform.",
    "ha2":"MDT Lung Cancer",
    "ha2d":"Weekly multidisciplinary oncology meeting.",
    "ha3":"Laboratory guides","ha3d":"Access to the laboratory guide.",
    "ha4":"Preoperative examinations",
    "ha4d":"Guidelines for preoperative evaluation.",
    "l_ct":"Contact &amp; appointment",
    "h_ct":"Book an appointment in Pulmonology",
    "h3_pr":"Practical information",
    "on_desc":"appointments can be booked 24/7 online via mynexuzhealth",
    "rook":"Smoking cessation clinic","dring":"Emergencies",
    "dring_d":"Pulmonologist on call 24/7 reachable via the AZORG emergency department",
    "h3_form":"Online request",
    "f_vn":"First name","f_an":"Last name","f_em":"Email address",
    "f_tel":"Phone number","f_camp":"Preferred campus",
    "f_camp_ph":"Select campus",
    "f_red":"Reason for consultation",
    "f_red_ph":"Select category",
    "f_info":"Additional information",
    "f_info_ph":"Describe your symptoms or mention your referring physician...",
    "f_submit":"Submit request →",
    "f_thx":"Thank you for your request.\\nOur secretariat will contact you within 2 working days.",
    "f_opts":["Shortness of breath / cough","Sleep apnoea / snoring",
              "Allergy / asthma","Smoking cessation",
              "Lung cancer follow-up","Pulmonary rehabilitation",
              "Sports clinic / diving fitness",
              "Existing condition check-up",
              "Referral from GP","Other"],
    "beh":"What we treat",
    "part":"Multidisciplinary collaboration with",
    "ft_tag":"Care from A to Z",
    "ft_desc":"Pulmonology is part of AZORG, the third largest hospital in Flanders.",
    "ft_aand":"Conditions","ft_klin":"Clinics","ft_pr":"Practical",
    "ft_on":"Book online","ft_dos":"My records",
    "ft_pat":"Patient info","ft_pro":"Professional info",
    "ft_copy":"© 2025 AZORG Hospital — Pulmonology / Respiratory Medicine",
  },
  "de": {
    "lang":"de",
    "page_title":"Pneumologie / Lungenheilkunde — AZORG Krankenhaus",
    "meta_desc":"Abteilung Pneumologie des AZORG. {n} Pneumologen an {c} Standorten.",
    "skip":"Zum Inhalt springen","back":"← AZORG Krankenhaus","news":"Nachrichten",
    "portal":"Patientenportal (mynexuzhealth)",
    "patient":"Patient","pro":"Fachkraft","praktisch":"Praktisch",
    "sub":"Lungenheilkunde",
    "n_over":"Über uns","n_camp":"Standorte","n_arts":"Ärzte",
    "n_aand":"Erkrankungen","n_klin":"Kliniken",
    "n_ha":"Überweisende Ärzte","n_contact":"Kontakt",
    "n_cta":"Termin buchen",
    "bc1":"AZORG","bc2":"Unsere medizinischen Abteilungen",
    "bc3":"Pneumologie / Lungenheilkunde",
    "pills":"Aktiv an","btn_afs":"Termin buchen",
    "btn_sec":"Sekretariate &amp; Standorte",
    "qa1":"Sekretariat anrufen","qa2":"Online-Termin",
    "qa2s":"über mynexuzhealth","qa3":"Erkrankungen",
    "qa3s":"Lungenerkrankungen &amp; Krankheitsbilder",
    "qa4":"Meine Akte","qa4s":"mynexuzhealth-Portal",
    "l_over":"Über die Abteilung","l_camp":"Sekretariat pro Standort",
    "h_camp":"Kontaktieren Sie uns an Ihrem nächsten Standort",
    "s_camp":"Vereinbaren Sie einen Termin telefonisch, per E-Mail oder online über mynexuzhealth. Alle Standorte sind montags bis freitags erreichbar.",
    "camp_arts":"Pneumologen an diesem Standort",
    "l_arts":"Unser Team","h_arts":"Spezialisierte Pneumologen",
    "s_arts":"Alle Pneumologen sind konventioniert. Jeder verfügt über eine breite Grundkompetenz, ergänzt durch Subspezialisierungen.",
    "conv":"✓ Konventioniert",
    "arts_link":"Profil &amp; Foto auf azorg.be",
    "btn_arts":"Alle Ärzte auf azorg.be ansehen →",
    "l_aand":"Lungenerkrankungen","h_aand":"Erkrankungen, die wir behandeln",
    "s_aand":'Weitere Informationen auf <a href="https://www.azorg.be/nl/longziekten" style="color:var(--red);font-weight:600" target="_blank" rel="noopener noreferrer">azorg.be/longziekten</a>.',
    "meer":"Mehr Info →",
    "l_ond":"Diagnostik &amp; Eingriffe",
    "h_ond":"Lungenuntersuchungen und -eingriffe",
    "s_ond":"Die Abteilung verfügt über moderne diagnostische und interventionelle Techniken.",
    "l_klin":"Spezialkliniken",
    "h_klin":"Expertkliniken innerhalb der Pneumologie",
    "s_klin":"Neben den allgemeinen Sprechstunden bietet die Abteilung spezialisierte Kliniken.",
    "l_nws":"Nachrichten","h_nws":"Aktuelles aus der Pneumologie",
    "s_nws":"Aktuelle Entwicklungen, Studien und Nachrichten.",
    "btn_nws":"Alle Nachrichten auf azorg.be ansehen →",
    "cta_btn":"Online-Termin vereinbaren →",
    "l_ha":"Für Fachkräfte","h_ha":"Informationen für überweisende Ärzte",
    "s_ha":"Schnelle Überweisung, digitale Konsultation und MDT-Besprechungen.",
    "ha1":"Nexuzhealth-Konsultation",
    "ha1d":"Digitale Konsultation über die sichere Nexuzhealth-Plattform.",
    "ha2":"MDT Lungenkrebs",
    "ha2d":"Wöchentliche multidisziplinäre onkologische Besprechung.",
    "ha3":"Laborleitfäden","ha3d":"Zugang zum Laborleitfaden.",
    "ha4":"Präoperative Untersuchungen",
    "ha4d":"Richtlinien für die präoperative Beurteilung.",
    "l_ct":"Kontakt &amp; Termin",
    "h_ct":"Termin in der Pneumologie vereinbaren",
    "h3_pr":"Praktische Informationen",
    "on_desc":"Termine können rund um die Uhr online über mynexuzhealth gebucht werden",
    "rook":"Raucherentwöhnungsklinik","dring":"Notfälle",
    "dring_d":"Pneumologe im Bereitschaftsdienst 24/7 über die Notaufnahme des AZORG erreichbar",
    "h3_form":"Online-Anfrage",
    "f_vn":"Vorname","f_an":"Nachname","f_em":"E-Mail-Adresse",
    "f_tel":"Telefonnummer","f_camp":"Bevorzugter Standort",
    "f_camp_ph":"Standort auswählen",
    "f_red":"Grund der Konsultation",
    "f_red_ph":"Kategorie auswählen",
    "f_info":"Zusätzliche Informationen",
    "f_info_ph":"Beschreiben Sie Ihre Beschwerden oder nennen Sie Ihren überweisenden Arzt...",
    "f_submit":"Anfrage absenden →",
    "f_thx":"Vielen Dank für Ihre Anfrage.\\nUnser Sekretariat wird sich innerhalb von 2 Werktagen bei Ihnen melden.",
    "f_opts":["Atemnot / Husten","Schlafapnoe / Schnarchen",
              "Allergie / Asthma","Raucherentwöhnung",
              "Lungenkrebs-Nachsorge","Lungenrehabilitation",
              "Sportklinik / Tauchtauglichkeit",
              "Kontrolle bestehender Erkrankung",
              "Überweisung vom Hausarzt","Sonstiges"],
    "beh":"Was wir behandeln",
    "part":"Multidisziplinäre Zusammenarbeit mit",
    "ft_tag":"Versorgung von A bis Z",
    "ft_desc":"Die Pneumologie ist Teil des AZORG, dem drittgrößten Krankenhaus Flanderns.",
    "ft_aand":"Erkrankungen","ft_klin":"Kliniken","ft_pr":"Praktisch",
    "ft_on":"Online-Termin","ft_dos":"Meine Akte",
    "ft_pat":"Patienteninfo","ft_pro":"Fachkräfte-Info",
    "ft_copy":"© 2025 AZORG Krankenhaus — Pneumologie / Lungenheilkunde",
  },
}



# ═════════════════════════════════════════════════════════════════════
# TRANSLATION ENGINE (multi-language)
# ═════════════════════════════════════════════════════════════════════

_translators = {}       # {lang: GoogleTranslator or False}
_cache = {}             # {"fr": {"text": "traduction"}, "en": {"text": "translation"}}
_cache_dirty = False


def _load_cache():
    global _cache
    if CACHE_FILE.exists():
        raw = json.loads(CACHE_FILE.read_text(encoding="utf-8"))
        # Support old flat format (auto-migrate to nested)
        if raw and not any(isinstance(v, dict) for v in raw.values()):
            _cache = {"fr": raw, "en": {}}
        else:
            _cache = raw
    for lang in ("fr", "en", "de"):
        _cache.setdefault(lang, {})


def _save_cache():
    if _cache_dirty:
        CACHE_FILE.write_text(
            json.dumps(_cache, ensure_ascii=False, indent=2), encoding="utf-8")


def _get_translator(lang):
    if lang not in _translators:
        try:
            from deep_translator import GoogleTranslator
            t = GoogleTranslator(source='nl', target=lang)
            t.translate("test")
            _translators[lang] = t
        except Exception:
            _translators[lang] = False
    return _translators[lang]


# Plaatsnamen die nooit vertaald mogen worden (pass-through)
NO_TRANSLATE = {
    "Aalst", "Asse", "Ninove", "Wetteren", "Geraardsbergen",
    "Aalst Moorselbaan", "Aalst Merestraat",
    "Aalst — Moorselbaan", "Aalst — Merestraat",
    "Campus ASZ", "Campus Asse", "Campus Geraardsbergen",
    "Campus Ninove", "Campus Wetteren",
    "Moorselbaan", "Merestraat",
}


def tr(text, lang):
    """Translate NL text to target lang. Cache → deep-translator → NL fallback."""
    if not text or not text.strip():
        return text

    stripped = text.strip()

    # Never translate place names
    if stripped in NO_TRANSLATE:
        return text

    if stripped.startswith(('http', 'mailto:', 'tel:', '+32', '0')) and ' ' not in stripped:
        return text
    if stripped.replace('.', '').replace(',', '').replace('+', '').isdigit():
        return text

    # Check cache
    if stripped in _cache.get(lang, {}):
        return _cache[lang][stripped]

    # Try deep-translator
    global _cache_dirty
    translator = _get_translator(lang)
    if translator:
        try:
            import re
            tags = {}
            clean = stripped
            for i, m in enumerate(re.finditer(r'<[^>]+>', stripped)):
                placeholder = f"__TAG{i}__"
                tags[placeholder] = m.group()
                clean = clean.replace(m.group(), placeholder, 1)

            result = translator.translate(clean)

            for placeholder, tag in tags.items():
                result = result.replace(placeholder, tag)

            _cache[lang][stripped] = result
            _cache_dirty = True
            print(f"    🌍 [{lang.upper()}] Vertaald: {stripped[:50]}...")
            return result
        except Exception as e:
            print(f"    ⚠️  [{lang.upper()}] Vertaling mislukt: {e}")

    print(f"    ⚠️  [{lang.upper()}] Geen vertaling voor: {stripped[:60]}...")
    return text


# ═════════════════════════════════════════════════════════════════════
# EXCEL READER
# ═════════════════════════════════════════════════════════════════════

def read_excel(filename):
    """Read Excel file from data/ directory, return list of dicts."""
    path = DATA_DIR / filename
    if not path.exists():
        print(f"  ⚠️  {filename} niet gevonden"); return []
    wb = openpyxl.load_workbook(path, data_only=True)
    sn = next((n for n in wb.sheetnames if n.upper() != "LEESMIJ"), wb.sheetnames[0])
    ws = wb[sn]
    headers = [c.value for c in ws[1] if c.value is not None]
    data = []
    for row in ws.iter_rows(min_row=2, max_col=len(headers), values_only=True):
        if any(v is not None for v in row):
            data.append({
                h: (str(row[i]).strip() if i < len(row) and row[i] is not None else "")
                for i, h in enumerate(headers)
            })
    return data


def read_teksten():
    """Read teksten.xlsx as {sectie: tekst} dict."""
    return {it["Sectie"]: it.get("Tekst", "")
            for it in read_excel("teksten.xlsx") if it.get("Sectie")}


def translate_items(items, fields, lang):
    """Translate specific fields of each item NL→lang."""
    result = []
    for item in items:
        loc = dict(item)
        for f in fields:
            if loc.get(f):
                loc[f] = tr(loc[f], lang)
        result.append(loc)
    return result


# ═════════════════════════════════════════════════════════════════════
# BUILD
# ═════════════════════════════════════════════════════════════════════

def build_lang(lang):
    ui = UI[lang]

    # Read teksten
    teksten_nl = read_teksten()
    if lang != "nl":
        t = {k: tr(v, lang) for k, v in teksten_nl.items()}
    else:
        t = teksten_nl

    # Read & sort data
    artsen_raw = sorted(
        [a for a in read_excel("artsen.xlsx") if a.get("Actief", "Ja") == "Ja"],
        key=lambda x: (x.get("Achternaam", ""), x.get("Voornaam", "")))

    campussen_raw = sorted(read_excel("campussen.xlsx"),
                           key=lambda x: int(x.get("Volgorde", 99)))
    aandoeningen_raw = sorted(read_excel("aandoeningen.xlsx"),
                              key=lambda x: int(x.get("Volgorde", 99)))
    klinieken_raw = sorted(read_excel("klinieken.xlsx"),
                           key=lambda x: int(x.get("Volgorde", 99)))
    nieuws_raw = [n for n in read_excel("nieuws.xlsx")
                  if n.get("Gepubliceerd", "Ja") == "Ja"]
    onderzoeken_raw = sorted(read_excel("onderzoeken.xlsx"),
                             key=lambda x: int(x.get("Volgorde", 99)))

    # Translate if not NL
    if lang != "nl":
        artsen = translate_items(artsen_raw, ["Functie", "Specialisatie"], lang)
        campussen = translate_items(campussen_raw,
                                    ["Naam", "Subtitel", "Openingsuren", "Extra info"], lang)
        aandoeningen = translate_items(aandoeningen_raw, ["Naam", "Beschrijving"], lang)
        klinieken = translate_items(klinieken_raw,
                                    ["Naam", "Beschrijving", "Link-tekst"], lang)
        nieuws = translate_items(nieuws_raw,
                                 ["Titel", "Categorie", "Samenvatting"], lang)
        onderzoeken = translate_items(onderzoeken_raw, ["Naam", "Beschrijving"], lang)
    else:
        artsen = artsen_raw
        campussen = campussen_raw
        aandoeningen = aandoeningen_raw
        klinieken = klinieken_raw
        nieuws = nieuws_raw
        onderzoeken = onderzoeken_raw

    # Link artsen to campussen (always NL names for doctors)
    for c in campussen:
        kort = c.get("Korte naam", "")
        c["_artsen"] = [a for a in artsen_raw
                        if a.get("Actief", "Ja") == "Ja"
                        and kort in [x.strip() for x in
                                     a.get("Campussen", "").split(",")]]

    # Render template
    tpl_text = (BASE / "template.html").read_text(encoding="utf-8")
    env = Environment(loader=BaseLoader(), autoescape=False)
    template = env.from_string(tpl_text)

    html = template.render(
        ui=ui, t=t,
        artsen=artsen, campussen=campussen, aandoeningen=aandoeningen,
        klinieken=klinieken, nieuws=nieuws, onderzoeken=onderzoeken,
        partners=PARTNERS[lang], cm=COLOR_MAP, ccc=CAMPUS_CC,
        gen=datetime.now().strftime("%d/%m/%Y %H:%M"),
        na=len(artsen),
    )

    # Write output
    OUTPUT_DIR.mkdir(exist_ok=True)
    SUFFIXES = {"nl": "", "fr": "-fr", "en": "-en", "de": "-de"}
    suffix = SUFFIXES.get(lang, f"-{lang}")
    outfile = OUTPUT_DIR / f"longziekten-azorg{suffix}.html"
    outfile.write_text(html, encoding="utf-8")
    print(f"  ✅ [{lang.upper()}] {outfile.name} "
          f"({len(html):,} tekens, {html.count(chr(10)):,} regels)")


def main():
    langs = ["nl", "fr", "en", "de"]
    if len(sys.argv) > 2 and sys.argv[1] == "--lang":
        langs = [sys.argv[2]]

    print("🔨 Website genereren...\n")
    _load_cache()
    total = sum(len(v) for v in _cache.values() if isinstance(v, dict))
    print(f"  📚 Vertaalcache: {total} vertalingen geladen\n")

    for lang in langs:
        build_lang(lang)

    _save_cache()
    if _cache_dirty:
        total = sum(len(v) for v in _cache.values() if isinstance(v, dict))
        print(f"\n  💾 Vertaalcache bijgewerkt: {total} vertalingen")

    print("\n🎉 Klaar!")


if __name__ == "__main__":
    main()
